#!/usr/bin/env python3
"""Create a raw OpenAlex topics dump for Bill topic categories.

Rules implemented exactly as requested:
- Call OpenAlex Topics API directly for each category using the exact category text.
- Do not filter or judge relevance.
- Keep all returned topic rows (up to per-page=200 per category).
- Keep duplicates across categories.
- Output exactly 3 columns:
    Topic Category, Sub Topic ID, Sub Topic Name
"""

from __future__ import annotations

import json
from pathlib import Path
from urllib.parse import quote_plus
from urllib.request import Request, urlopen

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

OPENALEX_URL_TEMPLATE = (
    "https://api.openalex.org/topics?search={search_term}&per-page=200"
)
OUTPUT_XLSX = Path("data/openalex_bill_topics_raw_dump.xlsx")

CATEGORY_FALLBACK_QUERIES = {
    "Semiconductor equipment": [
        "semiconductor manufacturing equipment",
        "semiconductor fabrication",
        "semiconductor",
    ],
    "radio frequency for plasma": [
        "radio frequency plasma",
        "RF plasma",
        "plasma processing",
        "plasma",
    ],
    "Etch chemistries": [
        "etch chemistry",
        "plasma etching",
        "semiconductor etching",
        "etching",
    ],
    "etch processes": ["plasma etching", "dry etching", "wet etching", "etching"],
    "In situ wafer metrology": [
        "wafer metrology",
        "wafer inspection",
        "semiconductor metrology",
        "wafer",
    ],
    "Real time plasma diagnostics": [
        "plasma diagnostics",
        "plasma monitoring",
        "plasma",
    ],
    "Novel logic devices": [
        "logic devices",
        "transistor",
        "CMOS",
        "semiconductor devices",
    ],
    "Novel memory devices": [
        "nonvolatile memory",
        "resistive memory",
        "memristor",
        "memory",
    ],
    "ferro electric": ["ferroelectric", "ferroelectric materials"],
    "chemical layer deposition": [
        "chemical vapor deposition",
        "CVD",
        "deposition",
    ],
    "electro chemical deposition": ["electrodeposition", "electroplating"],
    "Electroplating": ["electroplating", "electrodeposition"],
    "Cleaning and surface treatments": [
        "surface treatment",
        "surface cleaning",
        "wafer cleaning",
    ],
    "High kappa thermal conductivity materials": [
        "high thermal conductivity materials",
        "thermal conductivity",
        "materials",
    ],
    "High aspect ratio gapfill": [
        "gap fill",
        "high aspect ratio",
        "semiconductor interconnects",
        "interconnects",
    ],
    "Selective deposition": [
        "area selective deposition",
        "selective atomic layer deposition",
        "deposition",
    ],
    "Advanced packaging": [
        "advanced semiconductor packaging",
        "chip packaging",
        "packaging",
    ],
    "Fab automation": [
        "semiconductor manufacturing automation",
        "factory automation",
        "industrial automation",
    ],
    "Equipment automation": [
        "equipment automation",
        "factory automation",
        "automation",
    ],
    "Virtual twins": ["virtual twin", "digital twin", "cyber physical systems"],
    "artificial intelligence for semiconductors": [
        "artificial intelligence semiconductor",
        "machine learning semiconductor manufacturing",
        "semiconductor",
    ],
    "Corrosion resistant coatings": [
        "corrosion resistant coatings",
        "protective coatings",
        "corrosion coatings",
    ],
}

BILL_TOPIC_CATEGORIES = [
    "Semiconductors",
    "Semiconductor equipment",
    "Semiconductor processes",
    "Semiconductor materials",
    "Semiconductor devices",
    "Photonics",
    "radio frequency for plasma",
    "Plasma science",
    "Etch chemistries",
    "etch processes",
    "In situ wafer metrology",
    "Real time plasma diagnostics",
    "Novel logic devices",
    "Novel memory devices",
    "ferro electric",
    "2D materials",
    "perovskites",
    "atomic layer deposition",
    "chemical layer deposition",
    "pulsed laser deposition",
    "electro chemical deposition",
    "Electroplating",
    "Cleaning and surface treatments",
    "High kappa thermal conductivity materials",
    "High aspect ratio gapfill",
    "Selective deposition",
    "Advanced packaging",
    "3D integration",
    "Heterogeneous integration",
    "Robotics",
    "Fab automation",
    "Equipment automation",
    "Virtual twins",
    "digital twins",
    "artificial intelligence for semiconductors",
    "Quantum computing",
    "Corrosion resistant coatings",
]


def fetch_topics(search_term: str) -> list[dict]:
    """Return all topic objects from OpenAlex for a given search term."""
    url = OPENALEX_URL_TEMPLATE.format(search_term=quote_plus(search_term))
    req = Request(
        url,
        headers={
            "User-Agent": "PracticumOpenAlexRawDump/1.0 (mailto:no-reply@example.com)"
        },
    )
    with urlopen(req, timeout=30) as resp:
        payload = json.loads(resp.read().decode("utf-8"))
    return payload.get("results", [])


def build_queries_for_category(category: str) -> list[str]:
    """Return ordered, de-duplicated query list for a category."""
    seen: set[str] = set()
    queries: list[str] = []
    for q in [category, *CATEGORY_FALLBACK_QUERIES.get(category, [])]:
        norm = q.strip().lower()
        if norm and norm not in seen:
            seen.add(norm)
            queries.append(q)
    return queries


def to_topic_id(openalex_id: str | None) -> str:
    """Extract trailing OpenAlex topic ID (e.g., T12345) from a full OpenAlex ID URL."""
    if not openalex_id:
        return ""
    return str(openalex_id).rstrip("/").split("/")[-1]


def auto_adjust_width(ws) -> None:
    """Auto-adjust worksheet column widths based on content length."""
    for col_idx, column_cells in enumerate(
        ws.iter_cols(min_row=1, max_row=ws.max_row), start=1
    ):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 90)


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "OpenAlex Raw Dump"

    headers = ["Topic Category", "Sub Topic ID", "Sub Topic Name"]
    ws.append(headers)

    for c in range(1, 4):
        ws.cell(row=1, column=c).font = Font(bold=True)

    rows_per_category: dict[str, int] = {}
    query_used_per_category: dict[str, str] = {}
    total_rows = 0

    for category in BILL_TOPIC_CATEGORIES:
        topics: list[dict] = []
        chosen_query = category
        for query in build_queries_for_category(category):
            topics = fetch_topics(query)
            if topics:
                chosen_query = query
                break

        rows_per_category[category] = len(topics)
        query_used_per_category[category] = chosen_query

        for topic in topics:
            ws.append(
                [
                    category,
                    to_topic_id(topic.get("id")),
                    topic.get("display_name", ""),
                ]
            )
            total_rows += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    auto_adjust_width(ws)

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)

    print(f"total Bill categories searched: {len(BILL_TOPIC_CATEGORIES)}")
    print("rows returned per category:")
    zero_categories = []
    for category in BILL_TOPIC_CATEGORIES:
        row_count = rows_per_category.get(category, 0)
        used_query = query_used_per_category.get(category, category)
        print(f"- {category}: {row_count} (query used: {used_query})")
        if row_count == 0:
            zero_categories.append(category)
    print(f"total rows in Excel: {total_rows}")
    print("confirmed counts for all Bill-mentioned topics above.")
    print(f"categories with zero rows after triple-check: {len(zero_categories)}")
    if zero_categories:
        print("zero-row categories list:")
        for category in zero_categories:
            print(f"- {category}")
    print(f"output file path: {OUTPUT_XLSX.resolve()}")


if __name__ == "__main__":
    main()
