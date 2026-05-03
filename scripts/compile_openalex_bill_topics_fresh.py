#!/usr/bin/env python3
"""Build a fresh OpenAlex topic mapping for Bill's topic list.

This script only uses live OpenAlex Topics API responses and writes:
    data/openalex_bill_topics_fresh.xlsx

Columns:
    - Bill Topic Category
    - Topic ID
    - Topic Name
"""

from __future__ import annotations

import json
from pathlib import Path
from urllib.parse import quote_plus
from urllib.request import urlopen

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OPENALEX_TOPICS_URL = "https://api.openalex.org/topics?search={query}&per-page=200"

BILL_TOPICS = [
    "Semiconductors",
    "Semiconductor equipment",
    "Semiconductor processes",
    "Semiconductor materials",
    "Semiconductor devices",
    "Photonics",
    "radio frequency for plasma",
    "Plasma science",
    "Etch chemistries",
    "All etch processes",
    "In situ wafer metrology",
    "Real time plasma diagnostics",
    "Novel logic devices",
    "Novel memory devices",
    "ferro electric",
    "2D materials",
    "perovskites",
    "atomic layer deposition",
    "chemical layer deposition",
    "pulsed laser deposition",
    "electro chemical deposition",
    "Electroplating",
    "Cleaning and surface treatments",
    "High kappa thermal conductivity materials",
    "High aspect ratio gapfill",
    "Selective deposition",
    "Advanced packaging",
    "3D integration",
    "Heterogeneous integration",
    "Robotics",
    "Fab automation",
    "Equipment automation",
    "Virtual twins",
    "digital twins",
    "artificial intelligence for semiconductors",
    "Quantum computing",
    "Corrosion resistant coatings",
]

CATEGORY_QUERIES = {
    "Semiconductors": ["semiconductors", "microelectronics", "integrated circuits"],
    "Semiconductor equipment": [
        "semiconductor equipment",
        "wafer processing equipment",
        "lithography tools",
    ],
    "Semiconductor processes": [
        "semiconductor process",
        "wafer fabrication process",
        "microfabrication",
    ],
    "Semiconductor materials": [
        "semiconductor materials",
        "oxide semiconductors",
        "chalcogenide semiconductors",
    ],
    "Semiconductor devices": [
        "semiconductor devices",
        "transistor devices",
        "semiconductor detectors",
    ],
    "Photonics": ["photonics", "optical devices", "photonic crystals", "fiber optics"],
    "radio frequency for plasma": [
        "radio frequency plasma",
        "rf plasma",
        "microwave plasma",
    ],
    "Plasma science": ["plasma science", "low temperature plasma", "industrial plasma"],
    "Etch chemistries": [
        "etch chemistry",
        "plasma etch chemistry",
        "reactive ion etch chemistry",
    ],
    "All etch processes": [
        "etch process",
        "etching",
        "reactive ion etching",
        "dry etching",
    ],
    "In situ wafer metrology": [
        "in situ wafer metrology",
        "wafer metrology",
        "inline inspection",
    ],
    "Real time plasma diagnostics": [
        "real time plasma diagnostics",
        "plasma diagnostics",
        "plasma monitoring",
    ],
    "Novel logic devices": [
        "novel logic devices",
        "logic transistor",
        "beyond CMOS logic",
    ],
    "Novel memory devices": [
        "novel memory devices",
        "emerging memory devices",
        "phase change memory",
    ],
    "ferro electric": [
        "ferroelectric",
        "ferroelectric materials",
        "piezoelectric materials",
    ],
    "2D materials": ["2D materials", "graphene", "transition metal dichalcogenides"],
    "perovskites": ["perovskites", "perovskite materials"],
    "atomic layer deposition": ["atomic layer deposition", "ALD"],
    "chemical layer deposition": [
        "chemical vapor deposition",
        "chemical deposition",
        "CVD",
    ],
    "pulsed laser deposition": ["pulsed laser deposition", "PLD"],
    "electro chemical deposition": [
        "electro chemical deposition",
        "electrochemical deposition",
        "electrodeposition",
    ],
    "Electroplating": ["electroplating", "metal electroplating"],
    "Cleaning and surface treatments": [
        "surface cleaning",
        "surface treatment",
        "wafer cleaning",
    ],
    "High kappa thermal conductivity materials": [
        "high thermal conductivity materials",
        "high k dielectric",
        "thermal interface materials",
    ],
    "High aspect ratio gapfill": [
        "high aspect ratio gapfill",
        "trench fill",
        "via fill",
    ],
    "Selective deposition": [
        "selective deposition",
        "area selective deposition",
        "selective ALD",
    ],
    "Advanced packaging": [
        "advanced packaging",
        "chiplet packaging",
        "wafer level packaging",
    ],
    "3D integration": ["3D integration", "3D IC", "TSV"],
    "Heterogeneous integration": [
        "heterogeneous integration",
        "multi-die integration",
        "system-in-package",
    ],
    "Robotics": [
        "robotics",
        "robot manipulation",
        "mobile robots",
        "soft robotics",
        "teleoperation",
    ],
    "Fab automation": [
        "fab automation",
        "semiconductor manufacturing automation",
        "smart factory semiconductors",
    ],
    "Equipment automation": [
        "equipment automation",
        "tool automation",
        "robotic process automation",
    ],
    "Virtual twins": ["virtual twins", "virtual twin"],
    "digital twins": ["digital twins", "digital twin"],
    "artificial intelligence for semiconductors": [
        "artificial intelligence for semiconductors",
        "machine learning semiconductor process",
        "AI wafer manufacturing",
    ],
    "Quantum computing": [
        "quantum computing",
        "quantum algorithms",
        "quantum architecture",
        "quantum software",
    ],
    "Corrosion resistant coatings": [
        "corrosion resistant coatings",
        "anti-corrosion coating",
        "protective coatings",
    ],
}

CATEGORY_KEYWORDS = {
    "Semiconductors": [
        "semiconductor",
        "microelectronic",
        "integrated circuit",
        "chip",
        "wafer",
    ],
    "Semiconductor equipment": [
        "equipment",
        "tool",
        "lithography",
        "etch",
        "deposition",
    ],
    "Semiconductor processes": [
        "process",
        "processing",
        "fabrication",
        "manufacturing",
        "etching",
        "deposition",
    ],
    "Semiconductor materials": [
        "materials",
        "thin film",
        "oxide",
        "ga",
        "si",
        "chalcogenide",
        "interface",
    ],
    "Semiconductor devices": ["device", "transistor", "detector", "tft", "circuit"],
    "Photonics": ["photonic", "optic", "laser", "fiber", "metasurface", "grating"],
    "radio frequency for plasma": ["plasma", "radio frequency", "rf", "microwave"],
    "Plasma science": ["plasma", "discharge", "ionized"],
    "Etch chemistries": ["etch", "chemistr"],
    "All etch processes": ["etch", "etching", "reactive ion", "dry etch", "wet etch"],
    "In situ wafer metrology": [
        "metrology",
        "inspection",
        "wafer",
        "in situ",
        "inline",
    ],
    "Real time plasma diagnostics": [
        "plasma",
        "diagnostic",
        "monitoring",
        "real-time",
        "real time",
    ],
    "Novel logic devices": ["logic", "device", "beyond cmos"],
    "Novel memory devices": ["memory", "device", "phase change", "non-volatile"],
    "ferro electric": ["ferroelectric", "piezoelectric", "multiferroic"],
    "2D materials": ["2d", "graphene", "dichalcogenide", "monolayer"],
    "perovskites": ["perovskite"],
    "atomic layer deposition": ["atomic layer deposition", "ald"],
    "chemical layer deposition": [
        "chemical vapor deposition",
        "cvd",
        "chemical deposition",
    ],
    "pulsed laser deposition": ["pulsed laser deposition", "pld"],
    "electro chemical deposition": [
        "electrochemical deposition",
        "electro chemical deposition",
        "electrodeposition",
    ],
    "Electroplating": ["electroplating", "electroplate"],
    "Cleaning and surface treatments": [
        "cleaning",
        "surface treatment",
        "surface treatments",
        "surface modification",
    ],
    "High kappa thermal conductivity materials": [
        "thermal conductivity",
        "high k",
        "dielectric",
        "thermal",
    ],
    "High aspect ratio gapfill": [
        "high aspect ratio",
        "gapfill",
        "trench fill",
        "via fill",
    ],
    "Selective deposition": ["selective deposition", "area selective"],
    "Advanced packaging": ["packaging", "chiplet", "wafer level", "interposer"],
    "3D integration": ["3d integration", "3d ic", "tsv", "through silicon via"],
    "Heterogeneous integration": [
        "heterogeneous integration",
        "multi-die",
        "system-in-package",
        "sip",
    ],
    "Robotics": ["robot", "robotics", "manipulation", "teleoperation", "locomotion"],
    "Fab automation": ["fab", "automation", "smart factory", "industrial"],
    "Equipment automation": ["equipment", "automation", "tool automation", "rpa"],
    "Virtual twins": ["virtual twin"],
    "digital twins": ["digital twin", "digital twins"],
    "artificial intelligence for semiconductors": [
        "artificial intelligence",
        "ai",
        "machine learning",
        "neural",
        "semiconductor",
        "wafer",
        "fab",
    ],
    "Quantum computing": [
        "quantum",
        "qubit",
        "quantum computing",
        "quantum algorithm",
        "quantum software",
    ],
    "Corrosion resistant coatings": [
        "corrosion",
        "coating",
        "anti-corrosion",
        "protective",
    ],
}

EXCLUDE_KEYWORDS = {
    "medical",
    "surgery",
    "clinical",
    "cancer",
    "urology",
    "nephrology",
    "psychiatry",
    "mental health",
    "agriculture",
    "crop",
    "education",
    "teaching",
    "law",
    "legal",
    "wetland",
    "conservation",
    "hospital",
    "patient",
}

KEEP_ENGINEERING_KEYWORDS = {
    "semiconductor",
    "wafer",
    "chip",
    "plasma",
    "etch",
    "deposition",
    "materials",
    "device",
    "robot",
    "automation",
    "optic",
    "photon",
    "quantum",
    "coating",
}


def extract_topic_id(openalex_id: str) -> str:
    return openalex_id.rstrip("/").split("/")[-1]


def fetch_topics(search_term: str) -> list[dict]:
    url = OPENALEX_TOPICS_URL.format(query=quote_plus(search_term))
    with urlopen(url, timeout=45) as response:
        payload = json.loads(response.read().decode("utf-8"))
    return payload.get("results", [])


def score_topic(display_name: str, search_query: str, category: str) -> int:
    name = display_name.lower()
    query = search_query.lower()
    category_text = category.lower()

    # Exclude obvious irrelevant domains unless strongly engineering-oriented.
    if any(term in name for term in EXCLUDE_KEYWORDS):
        if not any(term in name for term in KEEP_ENGINEERING_KEYWORDS):
            return -999

    score = 0

    if query in name:
        score += 4
    if category_text in name:
        score += 4

    query_tokens = [
        token for token in query.replace("-", " ").split() if len(token) > 2
    ]
    score += sum(1 for token in query_tokens if token in name)

    category_keywords = CATEGORY_KEYWORDS.get(category, [])
    score += 2 * sum(1 for kw in category_keywords if kw in name)

    # Encourage semiconductor/engineering relevance where appropriate.
    if category != "Robotics" and any(
        k in category_text
        for k in ["semiconductor", "plasma", "etch", "deposition", "packaging"]
    ):
        if any(
            k in name
            for k in [
                "semiconductor",
                "wafer",
                "etch",
                "plasma",
                "deposition",
                "thin film",
                "fabrication",
                "chip",
            ]
        ):
            score += 2

    return score


def build_rows() -> tuple[list[tuple[str, str, str]], int]:
    rows: list[tuple[str, str, str]] = []
    no_match_count = 0

    for category in BILL_TOPICS:
        queries = CATEGORY_QUERIES.get(category, [category])
        best_by_topic: dict[str, tuple[str, int]] = {}

        for query in queries:
            try:
                results = fetch_topics(query)
            except Exception:
                continue

            for item in results:
                openalex_id = item.get("id")
                display_name = item.get("display_name")
                if not openalex_id or not display_name:
                    continue

                topic_id = extract_topic_id(openalex_id)
                score = score_topic(display_name, query, category)

                # Keep only strong-enough matches for this category.
                if score < 3:
                    continue

                existing = best_by_topic.get(topic_id)
                if existing is None or score > existing[1]:
                    best_by_topic[topic_id] = (display_name, score)

        if not best_by_topic:
            rows.append((category, "", "No clear OpenAlex topic found"))
            no_match_count += 1
            continue

        category_rows = [
            (category, topic_id, vals[0]) for topic_id, vals in best_by_topic.items()
        ]
        category_rows.sort(key=lambda r: r[2].lower())
        rows.extend(category_rows)

    # Sort final output by Bill Topic Category then Topic Name.
    rows.sort(key=lambda r: (r[0].lower(), r[2].lower()))
    return rows, no_match_count


def autosize_columns(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 80)


def apply_formatting(ws) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(
        min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            cell.alignment = Alignment(vertical="center")
            cell.border = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def main() -> None:
    repo_root = Path(__file__).resolve().parents[1]
    output_path = repo_root / "data" / "openalex_bill_topics_fresh.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    rows, no_match_count = build_rows()

    wb = Workbook()
    ws = wb.active
    ws.title = "OpenAlex Bill Topics"
    ws.append(["Bill Topic Category", "Topic ID", "Topic Name"])

    for row in rows:
        ws.append(list(row))

    apply_formatting(ws)
    autosize_columns(ws)
    wb.save(output_path)

    print(f"number of Bill topics searched: {len(BILL_TOPICS)}")
    print(f"number of categorized rows generated: {len(rows)}")
    print(f"number of Bill topics with no clear OpenAlex topic: {no_match_count}")
    print(f"output file path: {output_path.resolve()}")


if __name__ == "__main__":
    main()
