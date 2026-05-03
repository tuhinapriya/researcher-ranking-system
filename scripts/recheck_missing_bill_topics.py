"""
Re-check the 21 missing Bill Topic Categories against OpenAlex Topics API
using multiple search variations, and update the Excel file.
"""

import time
import requests
import openpyxl
from copy import copy

INPUT_FILE = "data/openalex_bill_topics_fresh.xlsx"
OUTPUT_FILE = "data/openalex_bill_topics_fresh_updated.xlsx"

MISSING_CATEGORIES = [
    "Advanced packaging",
    "All etch processes",
    "artificial intelligence for semiconductors",
    "atomic layer deposition",
    "chemical layer deposition",
    "Corrosion resistant coatings",
    "digital twins",
    "Electroplating",
    "Etch chemistries",
    "Fab automation",
    "Heterogeneous integration",
    "High aspect ratio gapfill",
    "In situ wafer metrology",
    "Novel logic devices",
    "Novel memory devices",
    "pulsed laser deposition",
    "radio frequency for plasma",
    "Selective deposition",
    "Semiconductor equipment",
    "Semiconductor processes",
    "Virtual twins",
]

# Multiple search variations per category (tries in order, stops at first hit)
SEARCH_VARIATIONS = {
    "Advanced packaging": [
        "advanced packaging",
        "advanced semiconductor packaging",
        "semiconductor packaging",
        "chip packaging",
        "heterogeneous packaging",
    ],
    "All etch processes": [
        "plasma etching",
        "semiconductor etching",
        "dry etching",
        "wet etching",
        "reactive ion etching",
    ],
    "artificial intelligence for semiconductors": [
        "artificial intelligence semiconductor",
        "machine learning semiconductor manufacturing",
        "AI chip design",
        "deep learning VLSI",
    ],
    "atomic layer deposition": [
        "atomic layer deposition",
        "ALD thin film",
        "atomic layer epitaxy",
    ],
    "chemical layer deposition": [
        "chemical vapor deposition",
        "CVD thin film",
        "PECVD",
        "LPCVD",
    ],
    "Corrosion resistant coatings": [
        "corrosion resistant coatings",
        "anti-corrosion coatings",
        "protective coatings",
        "corrosion protection",
    ],
    "digital twins": [
        "digital twin",
        "cyber physical systems",
        "digital twin manufacturing",
    ],
    "Electroplating": [
        "electroplating",
        "electrodeposition",
        "copper electroplating semiconductor",
    ],
    "Etch chemistries": [
        "etch chemistry",
        "plasma etch chemistry",
        "semiconductor etching chemistry",
        "fluorine etch",
    ],
    "Fab automation": [
        "semiconductor fab automation",
        "factory automation semiconductor",
        "industrial automation semiconductor",
        "automated semiconductor manufacturing",
    ],
    "Heterogeneous integration": [
        "heterogeneous integration",
        "chiplet integration",
        "2.5D 3D integration",
        "die integration",
    ],
    "High aspect ratio gapfill": [
        "high aspect ratio gap fill",
        "gap fill dielectric",
        "high aspect ratio trench fill",
        "semiconductor interconnect fill",
    ],
    "In situ wafer metrology": [
        "in situ wafer metrology",
        "wafer metrology",
        "semiconductor metrology",
        "wafer inspection",
        "process control metrology",
    ],
    "Novel logic devices": [
        "FinFET",
        "GAAFET",
        "nanosheet transistor",
        "logic devices transistor scaling",
        "beyond CMOS",
    ],
    "Novel memory devices": [
        "memristor",
        "resistive memory",
        "phase change memory",
        "nonvolatile memory",
        "MRAM",
    ],
    "pulsed laser deposition": [
        "pulsed laser deposition",
        "PLD thin film",
        "laser ablation deposition",
    ],
    "radio frequency for plasma": [
        "RF plasma",
        "radio frequency plasma",
        "plasma processing RF",
        "inductively coupled plasma",
    ],
    "Selective deposition": [
        "area selective deposition",
        "selective atomic layer deposition",
        "selective deposition",
        "surface selective deposition",
    ],
    "Semiconductor equipment": [
        "semiconductor manufacturing equipment",
        "semiconductor fabrication equipment",
        "wafer processing equipment",
        "semiconductor process equipment",
    ],
    "Semiconductor processes": [
        "semiconductor manufacturing process",
        "semiconductor fabrication process",
        "CMOS fabrication",
        "IC manufacturing",
    ],
    "Virtual twins": [
        "virtual twin",
        "digital twin simulation",
        "cyber physical twin",
        "virtual manufacturing model",
    ],
}

HEADERS = {"User-Agent": "researcher-kb-pipeline/1.0 (mailto:research@example.com)"}


def search_openalex_topics(term: str) -> list[dict]:
    """Query OpenAlex topics API and return list of {id, display_name, score}."""
    url = "https://api.openalex.org/topics"
    params = {"search": term, "per-page": 200}
    try:
        resp = requests.get(url, params=params, headers=HEADERS, timeout=20)
        resp.raise_for_status()
        data = resp.json()
        results = data.get("results", [])
        return [
            {
                "id": r.get("id", "").replace("https://openalex.org/", ""),
                "name": r.get("display_name", ""),
                "score": r.get("relevance_score", 0),
            }
            for r in results
        ]
    except Exception as e:
        print(f"  ERROR querying '{term}': {e}")
        return []


def best_topic_for_category(category: str) -> tuple[str, str] | tuple[None, None]:
    """
    Try each search variation for the category.
    Return (topic_id, topic_name) of the top result, or (None, None).
    """
    variations = SEARCH_VARIATIONS.get(category, [category])
    for term in variations:
        print(f"  Trying: '{term}'")
        results = search_openalex_topics(term)
        time.sleep(0.3)  # be polite to the API
        if results:
            top = results[0]
            print(
                f"    -> Found: {top['id']}  |  {top['name']}  (score={top['score']:.4f})"
            )
            return top["id"], top["name"]
    return None, None


def main():
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active

    # Identify header row columns
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    print(f"Columns: {headers}")

    # Find column indices (1-based)
    try:
        col_category = headers.index("Bill Topic Category") + 1
        col_id = headers.index("Topic ID") + 1
        col_name = headers.index("Topic Name") + 1
    except ValueError as e:
        raise RuntimeError(f"Expected column not found: {e}. Got: {headers}")

    # Normalised set of missing category names for matching
    missing_norm = {c.strip().lower(): c for c in MISSING_CATEGORIES}

    filled = []
    still_missing = []

    for row in ws.iter_rows(min_row=2):
        cat_cell = row[col_category - 1]
        id_cell = row[col_id - 1]
        name_cell = row[col_name - 1]

        category_val = str(cat_cell.value or "").strip()
        name_val = str(name_cell.value or "").strip()

        if "no clear" not in name_val.lower():
            continue  # already has a valid topic, skip

        cat_norm = category_val.lower()
        matched_key = None
        for k in missing_norm:
            if k in cat_norm or cat_norm in k:
                matched_key = missing_norm[k]
                break

        if matched_key is None:
            # Fallback: use the raw category value as search term
            matched_key = category_val

        print(f"\nChecking category: '{category_val}' (matched to '{matched_key}')")
        topic_id, topic_name = best_topic_for_category(matched_key)

        if topic_id:
            id_cell.value = topic_id
            name_cell.value = topic_name
            filled.append(category_val)
        else:
            id_cell.value = None
            name_cell.value = "No clear OpenAlex topic found"
            still_missing.append(category_val)

    wb.save(OUTPUT_FILE)
    print(f"\n{'='*60}")
    print(f"Output saved to: {OUTPUT_FILE}")
    print(f"\nSUMMARY")
    print(
        f"  Previously missing categories checked : {len(filled) + len(still_missing)}"
    )
    print(f"  Filled with at least one topic        : {len(filled)}")
    print(f"  Still missing                         : {len(still_missing)}")
    if still_missing:
        print("\nStill missing categories:")
        for c in still_missing:
            print(f"  - {c}")
    else:
        print("\nAll categories now have a topic!")


if __name__ == "__main__":
    main()
