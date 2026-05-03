#!/usr/bin/env python3
"""Expand Lam categories into OpenAlex topic candidates.

This script discovers OpenAlex topics per category using multiple search phrases,
then exports an Excel workbook with:
1) Detailed rows: Category, Topic ID, Topic Name
2) Category counts: Category, Number of Topics
"""

from __future__ import annotations

import json
import time
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Tuple
from urllib.parse import quote_plus
from urllib.request import urlopen

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

OPENALEX_TOPICS_URL = "https://api.openalex.org/topics?search={term}&per-page=200"

CATEGORY_CONFIG: Dict[str, Dict[str, object]] = {
    "Semiconductors": {
        "queries": [
            "semiconductor",
            "microelectronics",
            "integrated circuits",
            "wafer fabrication",
        ],
        "include": [
            "semiconductor",
            "microelectronic",
            "integrated circuit",
            "wafer",
            "chip",
        ],
        "exclude": ["education", "psychiatry", "wetland"],
        "min_score": 2,
    },
    "Semiconductor equipment": {
        "queries": [
            "semiconductor equipment",
            "wafer processing equipment",
            "lithography tool",
            "fab tools",
        ],
        "include": [
            "equipment",
            "tool",
            "lithography",
            "etch",
            "deposition",
            "fab",
            "wafer",
        ],
        "exclude": ["surgery", "nephrology"],
        "min_score": 2,
    },
    "Semiconductor processes": {
        "queries": [
            "semiconductor processing",
            "wafer process",
            "microfabrication process",
            "semiconductor manufacturing process",
        ],
        "include": ["process", "processing", "fabrication", "manufacturing", "wafer"],
        "min_score": 2,
    },
    "Semiconductor materials": {
        "queries": [
            "semiconductor materials",
            "thin films semiconductors",
            "oxide semiconductors",
            "chalcogenide semiconductor",
        ],
        "include": [
            "semiconductor",
            "materials",
            "thin film",
            "oxide",
            "gan",
            "sic",
            "chalcogenide",
            "ga2o3",
            "interface",
        ],
        "min_score": 2,
    },
    "Semiconductor devices": {
        "queries": [
            "semiconductor devices",
            "transistor technologies",
            "integrated circuit devices",
            "device physics semiconductor",
        ],
        "include": [
            "device",
            "transistor",
            "detector",
            "semiconductor",
            "circuit",
            "tft",
        ],
        "min_score": 2,
    },
    "Photonics": {
        "queries": [
            "photonics",
            "optical devices",
            "laser systems",
            "fiber optics",
            "photonic crystals",
        ],
        "include": [
            "photonic",
            "optic",
            "laser",
            "fiber",
            "metasurface",
            "metamaterial",
            "grating",
        ],
        "min_score": 2,
    },
    "RF radio frequency for plasma": {
        "queries": [
            "radio frequency plasma",
            "rf plasma",
            "microwave plasma",
            "plasma excitation RF",
        ],
        "include": ["plasma", "rf", "radio frequency", "microwave"],
        "min_score": 2,
    },
    "Plasma science": {
        "queries": [
            "plasma science",
            "low temperature plasma",
            "industrial plasma",
            "plasma processing",
        ],
        "include": ["plasma", "discharge", "ionized"],
        "min_score": 2,
    },
    "Etch chemistries": {
        "queries": [
            "etch chemistry",
            "plasma etch chemistry",
            "reactive ion etch chemistry",
        ],
        "include": ["etch", "chemistr", "reactive ion", "plasma"],
        "min_score": 2,
    },
    "All etch processes": {
        "queries": [
            "etch process",
            "reactive ion etching",
            "dry etching",
            "wet etching",
        ],
        "include": ["etch process", "etching", "reactive ion", "dry etch", "wet etch"],
        "min_score": 2,
    },
    "In-situ wafer metrology": {
        "queries": [
            "wafer metrology",
            "in situ metrology",
            "process metrology semiconductor",
            "inline wafer inspection",
        ],
        "include": [
            "metrology",
            "wafer",
            "inspection",
            "in situ",
            "inline",
            "defect detection",
        ],
        "min_score": 2,
    },
    "Real-time plasma diagnostics": {
        "queries": [
            "real-time plasma diagnostics",
            "plasma diagnostics",
            "in situ plasma monitoring",
        ],
        "include": [
            "plasma",
            "diagnostic",
            "monitoring",
            "real-time",
            "real time",
            "in situ",
        ],
        "min_score": 2,
    },
    "Novel logic devices": {
        "queries": [
            "novel logic devices",
            "logic transistor",
            "beyond CMOS logic",
            "quantum-dot cellular automata",
        ],
        "include": [
            "logic",
            "device",
            "transistor",
            "beyond cmos",
            "qca",
            "cellular automata",
        ],
        "min_score": 2,
    },
    "Novel memory devices": {
        "queries": [
            "novel memory devices",
            "emerging memory devices",
            "resistive memory",
            "phase change memory",
        ],
        "include": [
            "memory",
            "device",
            "phase change",
            "ferroelectric",
            "resistive",
            "non-volatile",
        ],
        "min_score": 2,
    },
    "Novel materials for electronics": {
        "queries": [
            "novel electronic materials",
            "2D materials electronics",
            "perovskite electronics",
            "ferroelectric materials",
            "graphene electronics",
        ],
        "include": [
            "graphene",
            "perovskite",
            "2d",
            "ferroelectric",
            "ga2o3",
            "chalcogenide",
            "oxide",
            "nanowire",
            "phase change",
            "multiferroic",
            "topological",
        ],
        "min_score": 1,
    },
    "ALD atomic layer deposition": {
        "queries": ["atomic layer deposition", "ALD thin films", "plasma enhanced ALD"],
        "include": ["atomic layer deposition", "ald"],
        "min_score": 2,
    },
    "CVD chemical vapor deposition": {
        "queries": ["chemical vapor deposition", "CVD", "plasma enhanced CVD"],
        "include": ["chemical vapor deposition", "cvd"],
        "min_score": 2,
    },
    "PLD pulsed laser deposition": {
        "queries": ["pulsed laser deposition", "PLD thin films"],
        "include": ["pulsed laser deposition", "pld"],
        "min_score": 2,
    },
    "ECD electro chemical deposition": {
        "queries": [
            "electrochemical deposition",
            "electro deposition",
            "electrodeposition thin films",
        ],
        "include": ["electrochemical deposition", "electrodeposition", "ecd"],
        "min_score": 2,
    },
    "Electroplating": {
        "queries": ["electroplating", "metal electroplating", "copper electroplating"],
        "include": ["electroplating", "electroplate"],
        "min_score": 2,
    },
    "Cleaning and surface treatments": {
        "queries": [
            "surface cleaning semiconductor",
            "surface treatment",
            "wafer cleaning",
        ],
        "include": [
            "cleaning",
            "surface treatment",
            "surface modification",
            "wafer cleaning",
        ],
        "min_score": 2,
    },
    "High kappa thermal conductivity materials": {
        "queries": [
            "high thermal conductivity materials",
            "high k dielectric",
            "thermal interface materials",
        ],
        "include": ["thermal conductivity", "high k", "dielectric", "thermal"],
        "min_score": 2,
    },
    "High-aspect ratio gapfill": {
        "queries": [
            "high aspect ratio gapfill",
            "gap fill process",
            "high aspect ratio trench fill",
        ],
        "include": ["high aspect ratio", "gapfill", "trench fill", "via fill"],
        "min_score": 2,
    },
    "Selective deposition": {
        "queries": [
            "selective deposition",
            "area selective deposition",
            "selective ALD",
        ],
        "include": ["selective deposition", "area selective", "selective ald"],
        "min_score": 2,
    },
    "Advanced packaging": {
        "queries": ["advanced packaging", "chiplet packaging", "wafer level packaging"],
        "include": ["packaging", "chiplet", "wafer level", "interposer"],
        "min_score": 2,
    },
    "3D integration": {
        "queries": ["3D integration", "3D IC", "TSV"],
        "include": ["3d integration", "3d ic", "tsv", "through silicon via"],
        "min_score": 2,
    },
    "Heterogeneous integration": {
        "queries": [
            "heterogeneous integration",
            "multi-die integration",
            "system-in-package",
        ],
        "include": [
            "heterogeneous integration",
            "multi-die",
            "system-in-package",
            "sip",
        ],
        "min_score": 2,
    },
    "Robotics": {
        "queries": [
            "robotics",
            "robot manipulation",
            "mobile robots",
            "soft robotics",
            "robotics control",
            "human robot interaction",
        ],
        "include": [
            "robotics",
            "robot",
            "hri",
            "manipulation",
            "locomotion",
            "teleoperation",
        ],
        "min_score": 1,
    },
    "Fab automation": {
        "queries": [
            "fab automation",
            "semiconductor manufacturing automation",
            "industrial automation fab",
        ],
        "include": [
            "fab",
            "automation",
            "manufacturing",
            "industrial",
            "line balancing",
        ],
        "min_score": 2,
    },
    "Equipment automation": {
        "queries": [
            "equipment automation",
            "tool automation",
            "robotic process automation",
        ],
        "include": [
            "equipment automation",
            "tool automation",
            "process automation",
            "rpa",
        ],
        "min_score": 2,
    },
    "Virtual twins or digital twins": {
        "queries": ["digital twins", "virtual twin", "physics-based digital twin"],
        "include": ["digital twin", "digital twins", "virtual twin"],
        "min_score": 2,
    },
    "AI artificial intelligence for semiconductors": {
        "queries": [
            "AI for semiconductors",
            "machine learning semiconductor process",
            "AI wafer manufacturing",
            "defect detection semiconductors",
        ],
        "include": [
            "ai",
            "artificial intelligence",
            "machine learning",
            "neural",
            "semiconductor",
            "wafer",
            "fab",
            "etch",
            "plasma",
        ],
        "min_score": 2,
    },
    "Quantum computing": {
        "queries": [
            "quantum computing",
            "quantum algorithms",
            "quantum architecture",
            "quantum software engineering",
        ],
        "include": ["quantum", "qubit", "quantum computing", "quantum algorithm"],
        "min_score": 2,
    },
    "Corrosion-resistant coatings": {
        "queries": [
            "corrosion-resistant coatings",
            "anti-corrosion coating",
            "protective coatings",
        ],
        "include": ["corrosion", "coating", "anti-corrosion", "protective"],
        "min_score": 2,
    },
}

CATEGORY_ORDER = list(CATEGORY_CONFIG.keys())


def extract_topic_id(openalex_id: str) -> str:
    return openalex_id.rstrip("/").split("/")[-1]


def fetch_topics(search_term: str) -> List[dict]:
    url = OPENALEX_TOPICS_URL.format(term=quote_plus(search_term))
    with urlopen(url, timeout=45) as response:
        payload = json.loads(response.read().decode("utf-8"))
    return payload.get("results", [])


def score_topic(
    name: str, query: str, include_terms: List[str], exclude_terms: List[str]
) -> int:
    n = name.lower()
    score = 0

    # Query phrase and token matching
    q = query.lower()
    if q in n:
        score += 4
    q_tokens = [token for token in q.replace("-", " ").split() if len(token) > 2]
    score += sum(1 for token in q_tokens if token in n)

    # Include/exclude keywords
    score += sum(1 for term in include_terms if term in n)
    score -= 2 * sum(1 for term in exclude_terms if term in n)

    # Semiconductor relevance boost for semiconductor-oriented categories
    if any(
        k in include_terms for k in ["semiconductor", "wafer", "fab", "etch", "plasma"]
    ):
        if any(
            k in n
            for k in [
                "semiconductor",
                "wafer",
                "chip",
                "fabrication",
                "etch",
                "deposition",
                "plasma",
                "thin-film",
                "thin film",
            ]
        ):
            score += 2

    return score


def discover_topics() -> Dict[str, Dict[str, str]]:
    # category -> topic_id -> topic_name
    per_category_candidates: Dict[str, Dict[str, Tuple[str, int]]] = {
        category: {} for category in CATEGORY_ORDER
    }

    for category in CATEGORY_ORDER:
        cfg = CATEGORY_CONFIG[category]
        raw_queries: List[str] = cfg["queries"]  # type: ignore[assignment]
        # Add the category label itself as an additional discovery query.
        queries: List[str] = list(dict.fromkeys(raw_queries + [category]))
        include_terms: List[str] = [s.lower() for s in cfg.get("include", [])]  # type: ignore[arg-type]
        exclude_terms: List[str] = [s.lower() for s in cfg.get("exclude", [])]  # type: ignore[arg-type]
        min_score: int = int(cfg.get("min_score", 1))

        for query in queries:
            try:
                results = fetch_topics(query)
            except Exception:
                continue

            for entry in results:
                openalex_id = entry.get("id")
                display_name = entry.get("display_name")
                if not openalex_id or not display_name:
                    continue

                topic_id = extract_topic_id(openalex_id)
                score = score_topic(display_name, query, include_terms, exclude_terms)
                if score < min_score:
                    continue

                current = per_category_candidates[category].get(topic_id)
                if current is None or score > current[1]:
                    per_category_candidates[category][topic_id] = (display_name, score)

            # polite pacing
            time.sleep(0.08)

    # Keep full per-category expansion and dedupe only within a category.
    per_category_final: Dict[str, Dict[str, str]] = {
        category: {} for category in CATEGORY_ORDER
    }
    for category in CATEGORY_ORDER:
        for topic_id, (topic_name, _score) in per_category_candidates[category].items():
            per_category_final[category][topic_id] = topic_name

    return per_category_final


def autosize_sheet_columns(ws, widths: Dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def apply_table_style(ws, max_col: int) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=max_col):
        for idx, cell in enumerate(row, start=1):
            cell.alignment = center if idx == 2 and max_col == 2 else left
            cell.border = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 22


def write_excel(
    per_category_topics: Dict[str, Dict[str, str]], output_path: Path
) -> None:
    wb = Workbook()

    ws_detail = wb.active
    ws_detail.title = "Expanded Topics"
    ws_detail.append(["Category", "Topic ID", "Topic Name"])

    for category in CATEGORY_ORDER:
        topics = per_category_topics.get(category, {})
        for topic_id, topic_name in sorted(topics.items(), key=lambda item: item[0]):
            ws_detail.append([category, topic_id, topic_name])

    apply_table_style(ws_detail, 3)
    autosize_sheet_columns(ws_detail, {"A": 44, "B": 14, "C": 68})

    ws_summary = wb.create_sheet("Category Counts")
    ws_summary.append(["Category", "Number of Topics"])

    for category in CATEGORY_ORDER:
        count = len(per_category_topics.get(category, {}))
        ws_summary.append([category, count])

    apply_table_style(ws_summary, 2)
    autosize_sheet_columns(ws_summary, {"A": 44, "B": 18})

    wb.save(output_path)


def main() -> None:
    repo_root = Path(__file__).resolve().parents[1]
    output_path = repo_root / "data" / "openalex_topics_expanded_by_category.xlsx"

    per_category_topics = discover_topics()
    write_excel(per_category_topics, output_path)

    total_topics = sum(len(v) for v in per_category_topics.values())
    print(str(output_path.resolve()))
    print(f"categories={len(CATEGORY_ORDER)}")
    print(f"total_topics={total_topics}")


if __name__ == "__main__":
    main()
